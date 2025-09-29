import os
import pandas as pd
from openpyxl import load_workbook

def save_results_to_xlsx(results, output_path, best_spec_dict):
    """
    Save results dictionary to an XLSX file where each key becomes a separate sheet.

    Args:
        results (dict): Dictionary where each key is a model name and value is a dict of specifications
        output_path (str): Path where the XLSX file should be saved
        best_spec (str): The best specification name to highlight in the output
    Returns:
        None
        Side effect: Creates/overwrites an XLSX file at output_path
    """
    #If sheet already exists, append to it, replacing sheets with the same name
    if os.path.exists(output_path):
        mode = "a"
        if_sheet_exists = "replace"

        # Ensure at least one sheet exists before writing
        wb = load_workbook(output_path)
        if len(wb.sheetnames) == 1:  # only one sheet, and it's about to be replaced
            # add a temporary sheet
            wb.create_sheet("TEMP_SHEET")
            wb.save(output_path)
    #Else, create a new file 
    else:
        mode = "w"
        if_sheet_exists = None
    # Create ExcelWriter object
    with pd.ExcelWriter(
        output_path, 
        engine="openpyxl",
        mode=mode,
        if_sheet_exists=if_sheet_exists
        ) as writer:
        # Iterate through each model in results
        for model_name, model_result in results.items():
            # If dictionary is empty we should skip saving the result
            if not model_result:
                continue
            df = pd.DataFrame()
            

            # Reorder columns to match desired format
            columns = [
                "Specification",
                "First Choice LL",
                "First Choice AIC",
                "Second Choice LL",
                "Second Choice RMSE",
                "Second Choice MAE",
                "Coefficient Names",
                "Estimated Coefficients",
                "Likelihood Ratio Test",
                "Predicted Diversion Matrix",
            ]
            # Add "Step" column if not "observables" model
            # There are no steps for the "observables" model
            if model_name != "observables":
                columns.append("Step")
                # Iterate through each step in the algorithm for each model (not "observables")
            df = pd.DataFrame.from_dict(model_result, orient="index")
            df.reset_index(inplace=True)
            df.rename(columns={"index": "Specification"}, inplace=True)
            
            df = df[columns]

            # Add "Best Specification" column to indicate the best specification 
            df["Best Specification"] = (df["Specification"] == best_spec_dict[model_name]).astype(int)

            # Write to Excel
            df.to_excel(
                writer,
                sheet_name=model_name,
                index=False,
            )

            # Get the worksheet to apply formatting
            worksheet = writer.sheets[model_name]

            # Format header
            for cell in worksheet[1]:
                cell.style = "Headline 2"

    # remove TEMP_SHEET if it was created
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        if "TEMP_SHEET" in wb.sheetnames:
            std = wb["TEMP_SHEET"]
            wb.remove(std)
            wb.save(output_path)


def save_predicted_diversion_matrices_to_xlsx(
    results, output_path, empirical_diversion_matrix, book_titles, best_spec_dict
):
    """
    Create a multi-sheet Excel file where each sheet contains the predicted diversion
    matrix (10 x 10) for the best model specification (lowest AIC) under each top-level
    key in 'results'. Also includes sheets for the empirical diversion matrix
    and the plain logit diversion matrix.
    Args:
        results (dict): Nested dictionary with model results
        output_path (str): Path to save the XLSX file
        empirical_diversion_matrix (list of list): 10x10 empirical diversion matrix
        book_titles (list of str): List of book titles corresponding to matrix rows/columns
        best_spec_dict (dict): Dictionary mapping model names to their best specification names
    Returns:
        None
        Side effect: Creates/overwrites n XLSX file at output_path with predicted diversion matrices for each model
    """
    # Make sure the directory for output_path already exists before writing
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    print(results.keys())

    #If sheet already exists, append to it, replacing sheets with the same name
    if os.path.exists(output_path):
        mode = "a"
        if_sheet_exists = "replace"
    #Else, create a new file 
    else:
        mode = "w"
        if_sheet_exists = None

    with pd.ExcelWriter(
        output_path, 
        engine="openpyxl",
        mode=mode,
        if_sheet_exists=if_sheet_exists
        ) as writer:
        # 1. Empirical Diversion Matrix
        df_empirical = pd.DataFrame(
            empirical_diversion_matrix, index=book_titles, columns=book_titles
        )
        df_empirical.to_excel(writer, sheet_name="data")

        # 2. Plain Logit Diversion Matrix (assumes 'plain logit' under 'observables')
        if "observables" in results and "plain logit" in results["observables"]:
            plain_logit_matrix = results["observables"]["plain logit"].get(
                "Predicted Diversion Matrix"
            )
            if plain_logit_matrix is not None:
                df_plain = pd.DataFrame(
                    plain_logit_matrix, index=book_titles, columns=book_titles
                )
                df_plain.to_excel(writer, sheet_name="plain_logit")

        # 3. For each top-level key in 'results', find the best specification by AIC
        for model_name, model_result in results.items():
            # If dictionary is empty we should skip saving the result
            if not model_result:
                continue

            ## If "observables", no steps; otherwise, iterate through steps
            selected_spec = model_result.get(best_spec_dict[model_name], None)
            if selected_spec is not None:
                best_pred_matrix = selected_spec.get(
                        "Predicted Diversion Matrix"
                    )
            df_pred = pd.DataFrame(
                best_pred_matrix, index=book_titles, columns=book_titles
            )
            sheet_name = f"{str(model_name)}"
            df_pred.to_excel(writer, sheet_name=sheet_name)