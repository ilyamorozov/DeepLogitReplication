# python -m src.replicate_comscore.4_comscore_estimation_mixed_logit
import os
import time
import warnings
import sys
import numpy as np
import pandas as pd
from scipy import stats
from openpyxl import load_workbook
import itertools

warnings.filterwarnings("ignore")

from src.helper_functions.estimation.estimate_mixed_logit import estimate_mixed_logit
from src.helper_functions.estimation.load_data import load_data_long_comscore
from src.helper_functions.file_structure.get_file_path_from_config import get_file_path_from_config

def format_spec_name(spec_list):
    if len(spec_list) == 1:
        return spec_list[0]
    elif len(spec_list) == 2:
        return f"{spec_list[0]} and {spec_list[1]}"
    else:
        return f"{', '.join(spec_list[:-1])}, and {spec_list[-1]}"

def gen_specs(k, embedding_model_name):
    """
    Returns dict:
      {0: {"plain logit with PCs": {}},
       1: { "<spec name>": {"price": "n"}, {"PC1": "n"}, {"PC2": "n"}, ...},
       ...
      }
    """
    pc_specifications = {}

    # Plain logit with PCs
    pc_specifications[0] = {"plain logit with PCs": {}}

    # Base parameter list: price + PC1 to PCk
    params = ["price"] + [f"PC{i}" for i in range(1, k + 1)]

    # Generate all non-empty combinations for r=1 to r=len(params)
    for r in range(1, len(params) + 1):
        pc_specifications[r] = {}  # Initialize dictionary for each r
        for combo in itertools.combinations(params, r):
            spec_name = format_spec_name(combo)
            pc_specifications[r][spec_name] = {
                f"{embedding_model_name}_{param.lower()}" 
                if param != "price" else param: "n"
                for param in combo
            }

    return pc_specifications



def save_results_to_xlsx(results, model_name, output_path):
    """
    Save results dict to an XLSX file. If the file exists, overwrite only the
    sheet named `model_name`; do not touch other sheets. If the file doesn't
    exist, create it.
    """
    # Build the DataFrame
    df = pd.DataFrame.from_dict(results, orient="index")
    df.index.name = "Specification"
    df.reset_index(inplace=True)

    # Keep only known columns if present
    columns = [
        "Specification",
        "First Choice LL",
        "First Choice AIC",
        "Coefficient Names",
        "Estimated Coefficients",
        "Standard Errors",
        "Likelihood Ratio Test",
    ]
    df = df[[c for c in columns if c in df.columns]]

    if os.path.exists(output_path):
        # Overwrite just the target sheet; keep other sheets intact
        try:
            with pd.ExcelWriter(
                output_path,
                engine="openpyxl",
                mode="a",
                if_sheet_exists="replace",
            ) as writer:
                df.to_excel(writer, sheet_name=model_name, index=False)
        except TypeError:
            # Fallback for very old pandas that doesn't support if_sheet_exists
            from openpyxl import load_workbook
            book = load_workbook(output_path)
            if model_name in book.sheetnames:
                book.remove(book[model_name])
                book.save(output_path)
            with pd.ExcelWriter(output_path, engine="openpyxl", mode="a") as writer:
                df.to_excel(writer, sheet_name=model_name, index=False)
    else:
        # Create new file with the single sheet
        with pd.ExcelWriter(output_path, engine="openpyxl", mode="w") as writer:
            df.to_excel(writer, sheet_name=model_name, index=False)



def get_best_aic(pc_specifications, results, r):

    best_spec = min(pc_specifications[r], key=lambda x: results[x]["First Choice AIC"]) if results else None
    best_aic = results[best_spec]["First Choice AIC"] if best_spec else None

    return best_spec, best_aic


def save_model_to_results(
    model,
    subresults,
    specification,
):
    if subresults == {}:
        subresults["plain logit"] = {
            "First Choice LL": model.loglikelihood,
            "First Choice AIC": model.aic,
            "Coefficient Names": model.coeff_names,
            "Estimated Coefficients": model.coeff_,
            "Standard Errors": model.stderr,
            "Likelihood Ratio Test": 0,
        }
        return

    best_model_subset = "plain logit"
    for model_subset in subresults:
        if (
            set(subresults[model_subset]["Coefficient Names"]).issubset(
                set(model.coeff_names)
            )
            and subresults[model_subset]["First Choice LL"]
            > subresults[best_model_subset]["First Choice LL"]
        ):
            best_model_subset = model_subset

    if model.loglikelihood < subresults[best_model_subset]["First Choice LL"]:
        model.loglikelihood = subresults[best_model_subset]["First Choice LL"]
        # Calculate AIC based on the new loglikelihood
        model.aic = -2 * model.loglikelihood + 2 * len(model.coeff_names)
        model.coeff_ = np.zeros(len(model.coeff_names))
        for i in range(len(subresults[best_model_subset]["Coefficient Names"])):
            model.coeff_[i] = subresults[best_model_subset]["Estimated Coefficients"][i]
    lr_test = (
        0
        if specification == "plain logit"
        else 2 * (model.loglikelihood - subresults["plain logit"]["First Choice LL"])
    )
    # number of random SD parameters which are additional parameters compared to baseline plain logit
    df = sum("sd" in name for name in model.coeff_names)
    
    result = {
        "First Choice LL": model.loglikelihood,
        "First Choice AIC": model.aic,
        "Coefficient Names": model.coeff_names,
        "Estimated Coefficients": model.coeff_,
        "Standard Errors": model.stderr,
        "Likelihood Ratio Test": stats.chi2.sf(lr_test, df) if df > 0 else 1.0,
    }
    subresults[specification] = result


def main():
    input_path = get_file_path_from_config(path_type="COMSCORE_4", path="INPUT_PATH")
    intermediate_path = get_file_path_from_config(path_type="COMSCORE_4", path="INTERMEDIATE_PATH")
    estimate_results_path = get_file_path_from_config(path_type="COMSCORE_4", path="RESULT_PATH")

    # categories = ["13060404", "13060114", "13060701", "13110101"]
    categories = pd.read_csv(os.path.join(input_path, "comscore_categories.csv"), dtype=str
        )["Category_Code"].tolist() # 40 categories

    task_index = int(os.getenv("SLURM_ARRAY_TASK_ID", sys.argv[1] if len(sys.argv)>1 else 0)) # Set array 0-39 in shell script
    category = categories[task_index]
    start_time = time.time()


    print(
        f"Estimating for category: {category}"
    )
    category_path = os.path.join(intermediate_path, category)
    first_choice_data = load_data_long_comscore(category_path)

    product_fixed_effects_varnames_minus_1 = sorted(
        [
            col
            for col in first_choice_data.columns
            if col.startswith("product_id_")
        ]
    )[:-1]

    embedding_models = sorted(
        list(
            set(
                [
                    col.split("_pc")[0]
                    for col in first_choice_data.columns
                    if "_pc" in col
                ]
            )
        )
    )

    # todays_date = time.strftime("%Y-%m-%d")
    todays_date = '2025-08-21'
    os.makedirs(estimate_results_path, exist_ok=True)
    output_path = os.path.join(
        estimate_results_path, f"mixed_logit_results_{category}_{todays_date}.xlsx"
    )

    models_reaching_max_PCs = []

    for em_idx, embedding_model in enumerate(embedding_models):
        print("=" * 40)
        print(
            f"Estimating for embedding model: {embedding_model}, {em_idx+1}/{len(embedding_models)}"
        )
    
        # Initialize results dictionary for this embedding model
        results = {}
        start_time = time.time()

        # NOTE: This model should include the full fixed effects (J-1)
        plain_logit_no_pc_model = estimate_mixed_logit(
            first_choice_data,
            product_fixed_effects_varnames_minus_1 + ["price"],
            {},
        )

        save_model_to_results(
            plain_logit_no_pc_model,
            results,
            "plain logit",
        )
        num_PCs = sum(col.startswith(f"{embedding_model}_pc") for col in first_choice_data.columns)
        k = min(6, num_PCs)

        if k == 0:
            print(f"No PCs found for {embedding_model}, skipping...")
            continue
        else:
            print(f"Found {num_PCs} PCs for {embedding_model}")

        pc_varnames = (
            product_fixed_effects_varnames_minus_1
            + ["price"]
            + [f"{embedding_model}_pc{i}" for i in range(1, k+1)]
        )

        plain_logit_with_pc_model = estimate_mixed_logit(
            first_choice_data,
            pc_varnames,
            {},
        )

        save_model_to_results(
            plain_logit_with_pc_model,
            results,
            "plain logit with PCs",
        )

        # assert likelihoods are close to equal
        if (
            plain_logit_with_pc_model.loglikelihood
            and plain_logit_no_pc_model.loglikelihood
        ):
            if (
                abs(
                    plain_logit_with_pc_model.loglikelihood
                    - plain_logit_no_pc_model.loglikelihood
                )
                > 1e-4
            ):
                print(
                    f"***Warning: Likelihoods are not close for plain logit with PCs and plain logit: {plain_logit_with_pc_model.loglikelihood} vs {plain_logit_no_pc_model.loglikelihood}"
                )

        pc_specifications = gen_specs(k, embedding_model)
        best_spec_so_far, best_aic_so_far = get_best_aic(pc_specifications, results, 0)
        tol = 1e-4

        # AIC stepwise over r
        for r in sorted(pc_specifications.keys()):
            print(f"----- Estimating models with {r} random variables -----")
            
            if r == 0:
                # already done above
                continue

            for specification, randvars in pc_specifications[r].items():
                try:
                    # Estimate the mixed logit model with the current specification
                    model = estimate_mixed_logit(
                        first_choice_data, pc_varnames, randvars
                    )

                    save_model_to_results(
                        model,
                        results,
                        specification,
                    )

                    print(f"Estimated specification: {specification}")
                except Exception as e:
                    print(f"Error for specification {specification}")
                    print(e)

                
            best_spec_this_r, best_aic_this_r = get_best_aic(pc_specifications, results, r) 
            
            # if there was an improvement in AIC compared to the best AIC for models with fewer random variables
            if best_aic_this_r + tol < best_aic_so_far: 
                print(f"  Improvement [{best_spec_so_far}:{best_aic_so_far:.4f}] -> [{best_spec_this_r}:{best_aic_this_r:.4f}]")  
                best_aic_so_far = best_aic_this_r
                best_spec_so_far = best_spec_this_r
                if r == k: 
                    models_reaching_max_PCs.append(embedding_model)
                 
            else: 
                print(f"  No improvement from AIC {best_aic_so_far:.4f}. Stopping here with best model: {best_spec_so_far} with K = {r-1}")
                break

            # save results for this embedding model as a new sheet in the output file
            save_results_to_xlsx(results, embedding_model, output_path) 

        end_time = time.time()
        elapsed_time = end_time - start_time
        print(f"Estimation complete for {embedding_model}, {elapsed_time:.2f} seconds")


    print("*" * 40)
    print("Models reaching max PCs:")
    print(models_reaching_max_PCs)

if __name__ == "__main__":
    main()


